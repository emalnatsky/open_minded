"""
Generate the CRI planning XLSX files requested by the supervisor.

Creates two files:
  1. um_with_errors_and_pregens.xlsx — Full UM + empty M1–M4 columns + empty L2-pregen columns
  2. um_with_errors_only.xlsx — Full UM + M1–M4 columns only (no utterances)

The column labels match exactly how the CRI script reads them from the API:
  - Mistake columns: M1_target_field, M1_wrong_value, M1_mistake_type, M1_spt_level, M1_step
  - Utterance columns: p1_hobby_bridge_comment, p1_m1_wrong_hobby_opener, etc.

USAGE:
    1. Start the FastAPI server: python main.py
    2. Run: python scripts/generate_cri_planning_xlsx.py
    3. Output goes to data/exports/

PREREQUISITES:
    pip install openpyxl requests
"""

import requests
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

API_URL = "http://localhost:8000"
OUTPUT_DIR = Path("data/exports")

# ═══════════════════════════════════════════════════════════════════════════════
# COLUMN DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════

# UM fields in display order (matches walkthrough flow)
UM_COLUMNS = [
    "child_id", "age", "exposure", "condition",
    # Hobby
    "hobbies", "hobby_fav", "hobby_talk",
    # Sport
    "sports_enjoys", "sports_fav", "sports_plays", "sports_fav_play",
    "sports_talk", "sports_play_talk",
    # Music
    "music_enjoys", "music_plays_instrument", "music_instrument", "music_talk",
    # Books
    "books_enjoys", "books_fav_genre", "books_fav_title", "books_talk",
    # Social
    "freetime_fav", "has_best_friend",
    # Animals
    "animals_enjoys", "animal_fav", "animal_talk",
    "has_pet", "pets", "pets_petName", "pet_talk",
    # Food
    "fav_food",
    # School
    "fav_subject", "school_strength", "school_difficulty",
    # Aspiration
    "interest", "aspiration", "role_model",
]

# Mistake columns — one set per M1–M4
MISTAKE_COLUMNS = []
for i in range(1, 5):
    MISTAKE_COLUMNS.extend([
        f"M{i}_target_field",
        f"M{i}_wrong_value",
        f"M{i}_mistake_type",
        f"M{i}_step",
    ])

# L2-pregen utterance step_ids — labels match how they're stored in GraphDB
# and how the CRI script reads them via GET /scenario
UTTERANCE_COLUMNS = [
    # Part 1: Hobby bridge
    "p1_hobby_bridge_comment",
    # Part 1: Topic 1 (generic — topic preselected per child)
    "topic_1",                              # metadata: "sport", "animals", "music", "books", or hobby fallback
    "p1_t1_recall",
    "p1_t1_open",
    "p1_t1_question",
    "p1_t1_followup",
    # Part 1: M1 mistake
    "p1_m1_wrong_hobby_opener",
    "p1_m1_followup_wrong_hobby",
    "p1_followup_postcorrection_true_hobby",
    # Part 1: Topic 2 (generic — topic preselected per child)
    "topic_2",                              # metadata: preselected second topic
    "p1_t2_open",
    "p1_t2_followup",
    "p1_t2_close",
    # Part 1: M2 mistake
    "p1_m2_followup_wrong_food",
    "p1_m2_postcorrection_true_food",
    # Part 2: School
    "p2_fav_subject_comment_subject",
    "p2_subject_profile_link",
    # Part 2: M3 mistake
    "p2_m3_postcorrection_true_strength",
    "p2_school_wrap_after_difficulty",
    # Part 3: Aspiration
    "p3_future_theme_wrap",
    "p3_rolemodel_recall",
    "p3_rolemodel_ack",
    "p3_norolemodel_ack",
    # Part 3: M4 mistake
    "p3_m4_followup_wrong_aspiration",
    "p3_m4_postcorrection_reflection",
    # Closing
    "closing",
]


# ═══════════════════════════════════════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════════════════════════════════════

FONT_HEADER = Font(name="Arial", bold=True, size=10, color="FFFFFF")
FONT_DATA = Font(name="Arial", size=10)
FONT_SECTION = Font(name="Arial", bold=True, size=10, color="FFFFFF")

FILL_UM = PatternFill("solid", fgColor="4472C4")         # blue — UM data
FILL_MISTAKE = PatternFill("solid", fgColor="C0504D")     # red — mistakes
FILL_UTTERANCE = PatternFill("solid", fgColor="9BBB59")   # green — utterances
FILL_ALT_ROW = PatternFill("solid", fgColor="F2F7FB")     # light blue alt rows

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_DATA = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="E0E0E0"),
)


# ═══════════════════════════════════════════════════════════════════════════════
# DATA FETCHING
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_all_children() -> list[dict]:
    """Fetch all children's flat profiles from the API."""
    resp = requests.get(f"{API_URL}/api/um/", timeout=10)
    resp.raise_for_status()
    child_ids = resp.json()["data"]["children"]

    rows = []
    for cid in child_ids:
        resp = requests.get(f"{API_URL}/api/um/{cid}/export", timeout=30)
        if resp.status_code != 200:
            print(f"  Skipping {cid}: {resp.status_code}")
            continue
        data = resp.json()
        current = data.get("current_profile", {})
        current["child_id"] = cid
        rows.append(current)
        print(f"  Fetched {cid}: {len([v for v in current.values() if v])} fields")

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# XLSX GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def write_xlsx(rows: list[dict], columns: list[str], col_fills: dict, filename: str):
    """Write an XLSX file with styled headers and data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CRI Planning"

    # ── Header row ──
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_HEADER

        # Color-code header by section
        if col_name in col_fills:
            cell.fill = col_fills[col_name]
        elif col_name.startswith("M") and "_" in col_name:
            cell.fill = FILL_MISTAKE
        elif col_name.startswith("p") or col_name == "closing":
            cell.fill = FILL_UTTERANCE
        else:
            cell.fill = FILL_UM

    # ── Data rows ──
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_DATA
            cell.alignment = ALIGN_DATA
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = FILL_ALT_ROW

    # ── Column widths ──
    for col_idx, col_name in enumerate(columns, 1):
        # UM columns narrower, utterance columns wider
        if col_name.startswith("p") or col_name == "closing":
            width = 45
        elif col_name.startswith("M") and "value" in col_name:
            width = 18
        else:
            max_len = max(len(col_name), max((len(str(r.get(col_name, ""))) for r in rows), default=5))
            width = min(max_len + 4, 30)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze and filter ──
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ──
    legend = wb.create_sheet("Legend")
    legend_data = [
        ("Color", "Section", "Description"),
        ("Blue", "UM Data", "Current child profile values from GraphDB"),
        ("Red", "Mistakes (M1–M4)", "Fill in: target_field, wrong_value, mistake_type, spt_level, step"),
        ("Green", "L2-pregen Utterances", "Fill in: the text Leo says at each step. Column name = step_id in GraphDB"),
        ("", "", ""),
        ("Mistake types:", "", "related-but-wrong | completely-wrong"),
        ("SPT levels:", "", "orientation | exploratory | affective"),
        ("Steps:", "", "4 (M1), 6 (M2), 10 (M3), 14 (M4)"),
        ("", "", ""),
        ("Utterance branches:", "", "Column = default branch. For corrected/not_corrected branches,"),
        ("", "", "add in the scenarios JSON after initial planning."),
    ]
    for r, (a, b, c) in enumerate(legend_data, 1):
        legend.cell(row=r, column=1, value=a).font = FONT_DATA
        legend.cell(row=r, column=2, value=b).font = FONT_DATA
        legend.cell(row=r, column=3, value=c).font = FONT_DATA
    legend.column_dimensions["A"].width = 15
    legend.column_dimensions["B"].width = 25
    legend.column_dimensions["C"].width = 70

    # Color legend cells
    legend.cell(row=2, column=1).fill = FILL_UM
    legend.cell(row=3, column=1).fill = FILL_MISTAKE
    legend.cell(row=4, column=1).fill = FILL_UTTERANCE

    # ── Save ──
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    print(f"\n  Saved: {filepath} ({len(rows)} children, {len(columns)} columns)")
    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("Fetching children from API...")
    rows = fetch_all_children()
    if not rows:
        print("No children found.")
        return

    print(f"\nFound {len(rows)} children.")

    # Build column fill map (all UM columns get blue fill)
    col_fills = {col: FILL_UM for col in UM_COLUMNS}

    # ── File 1: Full planning (UM + Mistakes + Utterances) ──
    print("\n--- Generating full planning XLSX ---")
    all_columns = UM_COLUMNS + MISTAKE_COLUMNS + UTTERANCE_COLUMNS
    write_xlsx(rows, all_columns, col_fills, "um_cri_planning_full.xlsx")

    # ── File 2: UM + Mistakes only (no utterances) ──
    print("\n--- Generating UM + mistakes only XLSX ---")
    mistakes_only_columns = UM_COLUMNS + MISTAKE_COLUMNS
    write_xlsx(rows, mistakes_only_columns, col_fills, "um_cri_planning_mistakes_only.xlsx")

    print("\nDone! Remove test/fake child rows manually in Excel before sharing.")


if __name__ == "__main__":
    main()