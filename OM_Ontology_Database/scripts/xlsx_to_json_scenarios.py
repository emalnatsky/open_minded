"""
Convert a filled-in CRI planning spreadsheet into scenarios JSON.

Reads the XLSX that the team filled in (UM values + mistakes + utterances)
and produces a scenarios JSON file ready for load_scenarios.py.

USAGE:
    python scripts/xlsx_to_scenarios.py data/cri_planning_filled.xlsx
    python scripts/xlsx_to_scenarios.py data/cri_planning_filled.xlsx --output data/scenarios.json

Then load into GraphDB:
    python scripts/load_scenarios.py data/scenarios_from_xlsx.json

WHAT IT DOES:
    For each row (child) in the spreadsheet:
    1. Reads M1–M4 mistake columns → builds mistake dicts
       - spt_level inferred: M1/M2=orientation, M3=exploratory, M4=affective
       - step defaults: M1=4, M2=6, M3=10, M4=14
    2. Reads utterance columns → builds utterance dicts
       - Branch is always "default"; the step_id name carries the meaning
         (for example postcorrection / followup_wrong)
       - Empty cells are skipped (no utterance created)
    3. Topic_1 and Topic_2 are stored as metadata utterances

SPREADSHEET FORMAT:
    The script expects the exact column headers from generate_cri_planning_xlsx.py,
    plus the team's filled-in values. UM columns (blue) are ignored during import.
"""

import json
import sys
import argparse
import logging
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

# Fixed values per walkthrough design
DEFAULT_STEPS = {"M1": 1.6, "M2": 1.8, "M3": 2.4, "M4": 3.4}
DEFAULT_SPT   = {"M1": "orientation", "M2": "orientation", "M3": "exploratory affective", "M4": "affective exchange"}

# Mistake column prefixes
MISTAKE_IDS = ["M1", "M2", "M3", "M4"]

# Columns that are UM reference values (not mistake data) — skip during import
UM_REFERENCE_COLS = {
    "child_id", "age", "exposure", "condition",
    "hobbies", "hobby_fav", "hobby_talk",
    "sports_enjoys", "sports_fav", "sports_plays", "sports_fav_play",
    "sports_talk", "sports_play_talk",
    "music_enjoys", "music_plays_instrument", "music_instrument", "music_talk",
    "books_enjoys", "books_fav_genre", "books_fav_title", "books_talk",
    "freetime_fav", "has_best_friend",
    "animals_enjoys", "animal_fav", "animal_talk",
    "has_pet", "pets", "pets_petName", "pet_talk",
    "fav_food", "fav_subject", "school_strength", "school_difficulty",
    "interest", "aspiration", "role_model",
}

# Utterance step_ids are stored under the default branch. The step_id name
# carries whether the text belongs to a corrected or not-corrected path.
def _infer_branch(step_id: str) -> str:
    """Return the branch used for CRI scenario utterances."""
    return "default"


# ═══════════════════════════════════════════════════════════════════════════════
# READER
# ═══════════════════════════════════════════════════════════════════════════════

def read_xlsx(filepath: str) -> list[dict]:
    """Read rows from XLSX. Returns list of dicts (one per child)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]

    data = []
    for row_vals in rows_iter:
        row = {}
        for h, v in zip(headers, row_vals):
            if h and v is not None:
                row[h] = str(v).strip()
        if row.get("child_id"):
            data.append(row)

    wb.close()
    logger.info("Read %d children from %s", len(data), filepath)
    return data


def read_csv(filepath: str) -> list[dict]:
    """Read rows from CSV. Returns list of dicts (one per child)."""
    import csv
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data = []
        for row in reader:
            cleaned = {k.strip(): v.strip() for k, v in row.items() if k and v and v.strip()}
            if cleaned.get("child_id"):
                data.append(cleaned)
    logger.info("Read %d children from %s", len(data), filepath)
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# CONVERTER
# ═══════════════════════════════════════════════════════════════════════════════

def row_to_scenario(row: dict) -> dict | None:
    """Convert one spreadsheet row into a scenario dict for load_scenarios.py."""
    child_id = row.get("child_id", "").strip()
    if not child_id:
        return None

    # ── Build mistakes ────────────────────────────────────────────────────
    mistakes = []
    for mid in MISTAKE_IDS:
        target = row.get(f"{mid}_target_field", "").strip()
        wrong  = row.get(f"{mid}_wrong_value", "").strip()
        mtype  = row.get(f"{mid}_mistake_type", "").strip()
        step   = row.get(f"{mid}_step", "").strip()

        if not target or not wrong:
            continue

        # Default mistake_type if empty
        if not mtype:
            mtype = "completely-wrong"

        # Default step from walkthrough structure
        if step:
            try:
                step = int(step)
            except ValueError:
                step = DEFAULT_STEPS[mid]
        else:
            step = DEFAULT_STEPS[mid]

        mistakes.append({
            "id": mid,
            "target_field": target,
            "wrong_value": wrong,
            "mistake_type": mtype,
            "spt_level": DEFAULT_SPT[mid],
            "step": step,
        })

    if not mistakes:
        logger.warning("  %s: no mistakes found, skipping", child_id)
        return None

    # ── Build utterances ──────────────────────────────────────────────────
    utterances = []

    for col_name, value in row.items():
        # Skip UM reference columns and mistake columns
        if col_name in UM_REFERENCE_COLS:
            continue
        if col_name.startswith(("M1_", "M2_", "M3_", "M4_")):
            continue
        if not value or not value.strip():
            continue

        step_id = col_name.strip()

        # Topic_1 and Topic_2 are metadata (topic label, not spoken text)
        if step_id in ("Topic_1", "topic_1"):
            utterances.append({
                "step_id": "topic_1",
                "layer": "L2-pregen",
                "branch": "default",
                "text": value.strip(),
            })
            continue
        if step_id in ("Topic_2", "topic_2"):
            utterances.append({
                "step_id": "topic_2",
                "layer": "L2-pregen",
                "branch": "default",
                "text": value.strip(),
            })
            continue

        # Regular utterance
        branch = _infer_branch(step_id)
        utterances.append({
            "step_id": step_id,
            "layer": "L2-pregen",
            "branch": branch,
            "text": value.strip(),
        })

    logger.info("  %s: %d mistakes, %d utterances", child_id, len(mistakes), len(utterances))

    return {
        "child_id": child_id,
        "mistakes": mistakes,
        "utterances": utterances,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Convert CRI planning spreadsheet to scenarios JSON."
    )
    parser.add_argument(
        "filepath",
        help="Path to the filled-in XLSX or CSV file"
    )
    parser.add_argument(
        "--output", "-o",
        help="Output JSON path (default: data/scenarios_from_xlsx.json)"
    )
    args = parser.parse_args()

    filepath = args.filepath
    ext = Path(filepath).suffix.lower()

    if ext in (".xlsx", ".xls"):
        rows = read_xlsx(filepath)
    elif ext in (".csv", ".tsv"):
        rows = read_csv(filepath)
    else:
        logger.error("Unsupported format: %s. Use .xlsx or .csv", ext)
        sys.exit(1)

    if not rows:
        logger.error("No data found.")
        sys.exit(1)

    # Convert each row
    scenarios = []
    for row in rows:
        scenario = row_to_scenario(row)
        if scenario:
            scenarios.append(scenario)

    if not scenarios:
        logger.error("No valid scenarios produced.")
        sys.exit(1)

    # Write JSON
    output_path = args.output or "data/scenarios_from_xlsx.json"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump({"scenarios": scenarios}, f, ensure_ascii=False, indent=2)

    logger.info("\nWrote %d scenarios to %s", len(scenarios), output_path)
    logger.info("Next: python scripts/load_scenarios.py %s", output_path)


if __name__ == "__main__":
    main()
