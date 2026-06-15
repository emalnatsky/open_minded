"""
Export child UM session data + CRI scenarios to JSON, CSV, or XLSX.

USAGE:
    python scripts/export_session.py child_001              # single child → JSON
    python scripts/export_session.py --all                   # all children → JSON
    python scripts/export_session.py --all --format csv      # all children → one CSV
    python scripts/export_session.py --all --format xlsx     # all children → one XLSX
    python scripts/export_session.py --all --format both     # all children → JSON + XLSX

The CSV/XLSX output creates a flat table with one row per child:
  - UM field values (blue headers)
  - M1–M4 mistake definitions (red headers)
  - Utterance texts by step_id (green headers)

Node fields with extra_props (like pets with petName) appear as companion
columns: e.g. "pets" → "Kat, Hond" and "pets_petName" → "donald en edgar, Rex".

PREREQUISITES:
    - The FastAPI server must be running (python main.py)
    - GraphDB must be running with the open-memory-robots repository
    - For XLSX: pip install openpyxl
"""

import sys
import csv
import json
import argparse
import requests
from pathlib import Path
from datetime import datetime, timezone

API_URL = "http://localhost:8000"

# Mistake column names (flat, per mistake)
MISTAKE_IDS = ["M1", "M2", "M3", "M4"]
MISTAKE_FIELDS = ["target_field", "wrong_value", "mistake_type", "spt_level", "step"]

# All utterance step_ids in script order
UTTERANCE_STEP_IDS = [
    "topic_1",
    "p1_hobby_bridge_comment",
    "p1_t1_recall", "p1_t1_open", "p1_t1_question", "p1_t1_followup",
    "p1_m1_wrong_hobby_opener", "p1_m1_followup_wrong_hobby",
    "topic_2",
    "p1_t2_open", "p1_t2_followup", "p1_t2_close",
    "p1_m2_followup_wrong_food", "p1_m2_postcorrection_true_food",
    "p2_fav_subject_comment_subject", "p2_subject_profile_link",
    "p2_m3_postcorrection_true_strength", "p2_school_wrap_after_difficulty",
    "p3_future_theme_wrap",
    "p3_rolemodel_recall", "p3_rolemodel_ack", "p3_norolemodel_ack",
    "p3_m4_followup_wrong_aspiration", "p3_m4_postcorrection_reflection",
    "closing",
]


# ═══════════════════════════════════════════════════════════════════════════════
# CRI SCENARIO FETCHING
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_scenario(child_id: str) -> dict | None:
    """Fetch CRI scenario for a child. Returns None if no scenario exists."""
    try:
        resp = requests.get(f"{API_URL}/api/um/{child_id}/scenario", timeout=15)
        if resp.status_code == 200:
            return resp.json()["data"]["scenario"]
        return None
    except Exception:
        return None


def flatten_scenario(scenario: dict | None) -> dict:
    """
    Flatten a CRI scenario into a dict with column-friendly keys.
    Returns: {
        "M1_target_field": "hobby_fav", "M1_wrong_value": "bakken", ...
        "p1_hobby_bridge_comment": "Dat is een leuke combinatie...", ...
    }
    """
    flat = {}

    if not scenario:
        return flat

    # Flatten mistakes
    mistakes_by_id = {}
    for m in scenario.get("mistakes", []):
        mistakes_by_id[m["id"]] = m

    for mid in MISTAKE_IDS:
        m = mistakes_by_id.get(mid, {})
        for field in MISTAKE_FIELDS:
            col = f"{mid}_{field}"
            flat[col] = str(m.get(field, "")) if m.get(field) is not None else ""

    # Flatten utterances (step_id → default branch text)
    utterances = scenario.get("utterances", {})
    for step_id in UTTERANCE_STEP_IDS:
        branches = utterances.get(step_id, {})
        # Take "default" branch; fall back to first available branch
        text = branches.get("default", "")
        if not text and branches:
            text = next(iter(branches.values()))
        flat[step_id] = text

    return flat


# ═══════════════════════════════════════════════════════════════════════════════
# JSON EXPORT (per-child, full detail)
# ═══════════════════════════════════════════════════════════════════════════════

def export_child_json(child_id: str, output_dir: str = "data/exports") -> Path | None:
    """Export a single child's full UM + history + CRI scenario to JSON."""
    resp = requests.get(f"{API_URL}/api/um/{child_id}/export", timeout=30)
    if resp.status_code == 404:
        print(f"  Skipping '{child_id}' — not found.")
        return None
    resp.raise_for_status()
    data = resp.json()

    # Add CRI scenario
    scenario = fetch_scenario(child_id)
    if scenario:
        data["cri_scenario"] = scenario
        n_mistakes = len(scenario.get("mistakes", []))
        n_utts = len(scenario.get("utterances", {}))
    else:
        data["cri_scenario"] = None
        n_mistakes = 0
        n_utts = 0

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = out_dir / f"{child_id}_export_{ts}.json"

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    s = data["summary"]
    print(f"  {child_id}: {s['total_fields_populated']} fields, "
          f"{s['total_change_events']} changes, "
          f"{n_mistakes} mistakes, {n_utts} utterances → {filename.name}")
    return filename


# ═══════════════════════════════════════════════════════════════════════════════
# FLAT TABLE EXPORT (CSV or XLSX)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_flat_rows(child_ids: list[str]) -> tuple[list[str], list[dict]]:
    """
    Fetch all children and build flat rows for tabular export.
    Returns (headers, rows).
    Includes UM values + CRI mistakes + CRI utterances.
    """
    schema_resp = requests.get(f"{API_URL}/api/schema", timeout=10)
    schema_resp.raise_for_status()
    all_fields = list(schema_resp.json()["data"]["fields"].keys())

    meta_columns = ["child_id", "exposure", "age", "created_at"]
    field_columns = [f for f in all_fields if f not in ("age", "exposure")]

    # Mistake columns
    mistake_columns = []
    for mid in MISTAKE_IDS:
        for field in MISTAKE_FIELDS:
            mistake_columns.append(f"{mid}_{field}")

    rows = []
    extra_columns_seen = set()

    for child_id in child_ids:
        resp = requests.get(f"{API_URL}/api/um/{child_id}/export", timeout=30)
        if resp.status_code != 200:
            print(f"  Skipping '{child_id}' — {resp.status_code}")
            continue

        data = resp.json()
        current = data.get("current_profile", {})
        full = data.get("full_profile", {})

        row = {}
        row["child_id"] = child_id
        row["exposure"] = current.get("exposure", "")
        row["age"] = current.get("age", "")
        row["created_at"] = full.get("scalars", {}).get("createdAt", {}).get("value", "")

        # Fill UM fields
        for field in field_columns:
            row[field] = current.get(field, "")

        # Pick up companion columns (e.g. pets_petName)
        for key, val in current.items():
            if key not in meta_columns and key not in field_columns:
                row[key] = val
                extra_columns_seen.add(key)

        # Fetch and flatten CRI scenario
        scenario = fetch_scenario(child_id)
        flat_cri = flatten_scenario(scenario)
        row.update(flat_cri)

        rows.append(row)

        n_um = len([v for k, v in row.items() if v and k in field_columns])
        has_cri = "yes" if scenario else "no"
        print(f"  {child_id}: {n_um} UM fields, CRI: {has_cri}")

    # Build final headers: UM + extras + mistakes + utterances
    extra_sorted = sorted(extra_columns_seen)
    headers = (meta_columns + field_columns + extra_sorted +
               mistake_columns + UTTERANCE_STEP_IDS)

    return headers, rows


def _get_column_section(col_name: str, meta_columns, field_columns, extra_columns,
                        mistake_columns) -> str:
    """Determine which section a column belongs to for color-coding."""
    if col_name in meta_columns or col_name in field_columns or col_name in extra_columns:
        return "um"
    if col_name in mistake_columns:
        return "mistake"
    return "utterance"


def export_all_csv(child_ids: list[str], output_dir: str = "data/exports") -> Path | None:
    """Export all children's UM + CRI data into one flat CSV."""
    if not child_ids:
        print("No children to export.")
        return None

    headers, rows = _build_flat_rows(child_ids)

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = out_dir / f"all_children_export_{ts}.csv"

    with open(filename, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    print(f"\n  CSV exported: {filename} ({len(rows)} children, {len(headers)} columns)")
    return filename


def export_all_xlsx(child_ids: list[str], output_dir: str = "data/exports") -> Path | None:
    """Export all children's UM + CRI data into one formatted XLSX."""
    if not child_ids:
        print("No children to export.")
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return None

    headers, rows = _build_flat_rows(child_ids)

    # Determine column sections for color-coding
    schema_resp = requests.get(f"{API_URL}/api/schema", timeout=10)
    all_fields = list(schema_resp.json()["data"]["fields"].keys())
    meta_columns = {"child_id", "exposure", "age", "created_at"}
    field_columns = set(f for f in all_fields if f not in ("age", "exposure"))
    extra_columns = set(h for h in headers if h not in meta_columns and h not in field_columns
                        and not h.startswith("M") and h not in UTTERANCE_STEP_IDS)
    mistake_columns = set(f"{mid}_{f}" for mid in MISTAKE_IDS for f in MISTAKE_FIELDS)

    wb = Workbook()
    ws = wb.active
    ws.title = "UM + CRI Export"

    # Styling
    font_header = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    fill_um = PatternFill("solid", fgColor="4472C4")
    fill_mistake = PatternFill("solid", fgColor="C0504D")
    fill_utterance = PatternFill("solid", fgColor="9BBB59")
    fill_alt = PatternFill("solid", fgColor="F2F7FB")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_data = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="E0E0E0"),
    )

    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_header

        section = _get_column_section(h, meta_columns, field_columns, extra_columns,
                                       mistake_columns)
        if section == "mistake":
            cell.fill = fill_mistake
        elif section == "utterance":
            cell.fill = fill_utterance
        else:
            cell.fill = fill_um

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(h, ""))
            cell.font = font_data
            cell.alignment = align_data
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = fill_alt

    # Column widths
    for col_idx, h in enumerate(headers, 1):
        if h in UTTERANCE_STEP_IDS:
            width = 50
        elif h in mistake_columns:
            width = 18
        else:
            max_len = len(h)
            for row_data in rows:
                val = str(row_data.get(h, ""))
                max_len = max(max_len, len(val))
            width = min(max_len + 4, 35)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend_rows = [
        ("Color", "Section", "Description"),
        ("Blue", "UM Data", "Current child profile values from GraphDB"),
        ("Red", "Mistakes (M1-M4)", "Deliberate wrong values for CRI interaction"),
        ("Green", "Utterances", "L2-pregen text Leo speaks at each step"),
    ]
    for r, (a, b, c) in enumerate(legend_rows, 1):
        legend.cell(row=r, column=1, value=a).font = font_data
        legend.cell(row=r, column=2, value=b).font = font_data
        legend.cell(row=r, column=3, value=c).font = font_data
    legend.cell(row=2, column=1).fill = fill_um
    legend.cell(row=3, column=1).fill = fill_mistake
    legend.cell(row=4, column=1).fill = fill_utterance
    legend.column_dimensions["A"].width = 15
    legend.column_dimensions["B"].width = 25
    legend.column_dimensions["C"].width = 60

    # Save
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = out_dir / f"all_children_export_{ts}.xlsx"
    wb.save(str(filename))

    print(f"\n  XLSX exported: {filename} ({len(rows)} children, {len(headers)} columns)")
    return filename


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_child_ids() -> list[str]:
    """Fetch all child IDs from the API."""
    resp = requests.get(f"{API_URL}/api/um/", timeout=10)
    resp.raise_for_status()
    return resp.json()["data"]["children"]


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export child UM + CRI scenario data to JSON, CSV, and/or XLSX."
    )
    parser.add_argument(
        "child_id", nargs="?",
        help="Child ID to export (omit if using --all)"
    )
    parser.add_argument(
        "--out", default="data/exports",
        help="Output directory (default: data/exports)"
    )
    parser.add_argument(
        "--all", action="store_true",
        help="Export every child in the database"
    )
    parser.add_argument(
        "--format", choices=["json", "csv", "xlsx", "both"], default="json",
        help="Output format: json, csv, xlsx, both (json + xlsx)"
    )
    args = parser.parse_args()

    if args.all:
        ids = get_all_child_ids()
        if not ids:
            print("No children found in GraphDB.")
            sys.exit(0)

        if args.format in ("json", "both"):
            print(f"\nExporting {len(ids)} children as JSON to {args.out}/")
            for cid in ids:
                export_child_json(cid, args.out)

        if args.format == "csv":
            print(f"\nExporting {len(ids)} children as CSV to {args.out}/")
            export_all_csv(ids, args.out)

        if args.format in ("xlsx", "both"):
            print(f"\nExporting {len(ids)} children as XLSX to {args.out}/")
            export_all_xlsx(ids, args.out)

        print("\nDone.")

    elif args.child_id:
        if args.format in ("csv", "xlsx"):
            print(f"{args.format.upper()} format requires --all (one row per child).")
            print("Use --format json for single child, or --all --format xlsx for all.")
            sys.exit(1)
        export_child_json(args.child_id, args.out)

    else:
        parser.print_help()
        sys.exit(1)